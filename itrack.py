#! python3

import datetime
import logging
import os
import sys
import time
import traceback
from contextlib import suppress
import threading
from typing import List, Dict, Mapping, Tuple, Union, Optional, Any
import random
import re
import collections
import pprint

import cv2
import numpy as np
from pyzbar.pyzbar import decode
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import pyinputplus
import pyinputplus as pyip

from util import getExcelFiles, getData, removeFromList

print(""" __      __  __     __                         ___               
/ _  /\ |__)(_ |__|(_   | _    _  _ |_ _  _     |  _ _  _ |  _  _
\__)/--\| \ __)|  |__)  || )\/(-`| )|_(_)| \/   | | (_|(_ |<(-`| 
                                           /                     
""")

defaultExcelFilename = "GARSHS_INVENTORY.xlsx"
itemColHead = "Item"
infos     = ["Available", "Place", "Person In Charge"]
tabNumber = [2, 3, 1]

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
# logging.disable(logging.INFO)

while True:
    excelFiles = getExcelFiles()
    name = input("Please input your name:")
    lastName = input("Last name:")
    fullName = name + " " + lastName


    noInitialFile = False

    if len(excelFiles) == 0:
        noInitialFile = True
        print('No xlsx files found in the directory.')
        print(
            f'Please enter a filename for the excel file to be created. Press enter if you would like the default name [{defaultExcelFilename}]')
        excelFilename = pyinputplus.inputStr(prompt='> ', blank=True)
        if excelFilename == '':
            excelFilename = defaultExcelFilename
        if not excelFilename.endswith('.xlsx'):
            excelFilename += '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(excelFilename)
        break
    elif len(excelFiles) == 1:
        excelFilename = excelFiles[0]
        print(f'Found {excelFilename}')
        print('Exit the program if this is not the excel file you want.')
        break
    elif len(excelFiles) == 2 and \
            (excelFiles[0].startswith('~$') or excelFiles[1].startswith('~$')):
        print('!!! The excel file is open in another window !!!')
        print('Please close the excel file then press enter to continue.')
        continueProgram = input('> ')
        continue
    else:
        exelFileIsOpen = False
        for file in os.listdir('.'):
            if exelFileIsOpen:
                break
            if file.startswith('~$') and file.endswith('.xlsx'):
                print('!!! An excel file is open in another window !!!')
                print('Please close the excel file/s then press enter to continue.')
                continueProgram = input('> ')
                for filename in os.listdir('.'):
                    if filename.startswith('~$') and file.endswith('.xlsx'):
                        exelFileIsOpen = True
                        break
        if exelFileIsOpen:
            continue
        print('Multiple xlsx files found in the directory.')
        excelFilename: str = pyinputplus.inputMenu(getExcelFiles(), numbered=True)
        break

# Detects if the excel file is not initialized (data is not yet inputted)
# and if not, write the column infos, and then prompts the user to fill it up
while True:
    wb = openpyxl.load_workbook(excelFilename)
    ws = wb.active
    rowsDict, columnsDict = getData(ws)
    maxRow = ws.max_row
    if not rowsDict:
        ws.freeze_panes = 'B3'
        print("Adding infos...")

        if ws['A1'].value != '':
            ws['A1'].value = 'GARSHS INVENTORY'
            ws['A1'].font = Font(size=20)
        if ws['A2'].value != '':
            ws['A2'].value = itemColHead
            ws['A2'].font = Font(size=14)
        for i in range(1, len(infos) + 1):
            if ws.cell(row=2, column=i+1).value != '':
                ws.cell(row=2, column=i+1).value = infos[i - 1]
                ws.cell(row=2, column=i+1).font = Font(size=14)

    try:
        wb.save(excelFilename)
    except PermissionError:
        print("\n!!!  The excel file is open on another window  !!!")
        print("Close the excel file to let the program manipulate it. Press enter if done closing the window.\n")
        continueProgram = input('> ')
        continue
    else:
        if maxRow <= 2:
            print("Please input at least one record in the excel file for the program to work.\n")
            print("Program ended.")
            sys.exit()
    break

# Puts the data to a dictionary
inventory = {}
for rowIndex in range(3, maxRow + 1):
    item = rowsDict[rowIndex][0]
    inventory.setdefault(item, {})
    for i in range(len(infos)):
        inventory[item][infos[i]] = rowsDict[rowIndex][i + 1]
        rowsDict[rowIndex][i + int(item)]  = inventory[item]

try:
    print("\nOpening the webcam... please wait\n")
    cap = cv2.VideoCapture(0)
    cap.set(3, 640)
    cap.set(4, 480)

    print("Webcam is running.     You can now show the QR codes to the webcam.")
    print("----------------------------ITEM INFO------------------------------")

    showedItems = []
    while True:
        success, img = cap.read()
        for qrcode in decode(img):
            item = qrcode.data.decode("utf-8")

            pts = np.array([qrcode.polygon], np.int32)
            pts = pts.reshape((-1, 1, 2))
            cv2.polylines(img, [pts], True, (255, 0, 255), 5)
            pts2 = qrcode.rect

            if item in inventory:
                cv2.putText(img, item, (pts2[0], pts2[1]), cv2.FONT_HERSHEY_COMPLEX, 0.9, (255, 0, 255), 2)
                if item not in showedItems:
                    print(f"{itemColHead}:\t\t\t{item}")
                    for i, info in enumerate(infos):
                        tabs = "\t" * tabNumber[i]
                        print(f"{info}:{tabs}{inventory[item][info]}")
                    print()
                    showedItems.append(item)
                    threading.Thread(target=removeFromList, args=[showedItems, item, 15]).start()
            else:
                cv2.putText(img, "UNRECOGNIZED", (pts2[0], pts2[1]), cv2.FONT_HERSHEY_COMPLEX, 0.9, (255, 0, 255), 2)

        cv2.imshow("Please show your QR code to the webcam.", img)
        cv2.waitKey(1)
except (KeyboardInterrupt, UserWarning):
    print("Program ended.")













